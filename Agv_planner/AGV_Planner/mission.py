"""
Env 2D
@author: ZiQi Qin
"""
import math

'''
包含2D地图信息，长宽，以及静态障碍物信息
包含车辆的起点和终点
'''
import random
import openpyxl
from create_map import get_map_data
from typing import List, Tuple, Dict, Callable, Set, Any

class Mission:
    def __init__(self, starts: List[Tuple[int, int]], goals: List[Tuple[int, int]]):
        self.env = Env(use_benchmark=True)
        self.x_range = self.env.x_range
        self.y_range = self.env.y_range
        if len(starts) != len(goals):
            raise ValueError("The number of starts and goals should be the same.")
        elif len(starts) == 0:
            self.mission_num = 5
            self.starts, self.goals = self.generate_random_mission(self.mission_num)
            print("Generate random mission successfully")
        else:
            self.starts = starts
            self.goals = goals
            self.mission_num = len(starts)
            self.check_mission_point(self.starts)
            self.check_mission_point(self.goals)

    def check_mission_point(self, points):
        for point in points:
            i = points.index(point)
            if point in self.env.obs:
                raise ValueError("mission point", i, point, " in obstacle")
            if point[0] < 0 or point[0] > self.x_range or point[1] < 0 or point[1] > self.y_range:
                raise ValueError("mission point", i, point, " out of range")

    def generate_random_mission(self, mission_num):
        starts = []
        goals = []
        num = 0
        while num == mission_num:
            start_x = random.randint(1, self.x_range-2)
            start_y = random.randint(1, self.y_range-2)
            goal_x = random.randint(1, self.x_range-2)
            goal_y = random.randint(1, self.y_range-2)
            start = (start_x, start_y)
            goal = (goal_x, goal_y)
            if start in self.env.obs or goal in self.env.obs:
                continue
            starts.append(start)
            goals.append(goal)
            num += 1
        return starts, goals

class Env:
    def __init__(self, use_benchmark=False):
        if use_benchmark:
            self.y_range, self.x_range, self.obs = get_map_data()
            # self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1)]  # 4 neighborhoods
            # self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1), (-1, 1), (1, 1), (-1, -1), (-1, 1)]  # 8 neighborhoods
        else:
            self.regenerate_obs = False
            self.x_range = 51  # size of background
            self.y_range = 31
            self.obs = self.generate_obs()
            # self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1)] # 4 neighborhoods
            # self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1), (-1, 1), (1, 1), (-1, -1), (-1, 1)]  # 8 neighborhoods


    def generate_obs(self):
        """
        Initialize obstacles' positions
        :return: map of obstacles
        """

        x = self.x_range
        y = self.y_range
        obs = set()
        # 地图边界
        for i in range(x):
            obs.add((i, 0))
        for i in range(x):
            obs.add((i, y - 1))
        for i in range(y):
            obs.add((0, i))
        for i in range(y):
            obs.add((x - 1, i))

        # 随机生成一系列障碍物
        if self.regenerate_obs:
            proportion = 0.2
            num_obstacles = int(proportion * (x - 2) * (y - 2))
            obstacle_data = []  # 存储障碍物坐标信息

            for _ in range(num_obstacles):
                rand_x = random.randint(1, x - 2)  # 排除边界
                rand_y = random.randint(1, y - 2)  # 排除边界
                obs.add((rand_x, rand_y))
                obstacle_data.append((rand_x, rand_y))
                self.save_obstacle_data_to_excel(obstacle_data)

            print("Generate obstacle data successfully")
            return obs
        return self.load_obs_from_excel(obs)

    def save_obstacle_data_to_excel(self, data):
        # 创建一个新的 Excel 文件
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # 添加表头
        sheet['A1'] = 'X Coordinate'
        sheet['B1'] = 'Y Coordinate'
        # 写入障碍物数据
        for index, (x, y) in enumerate(data, start=2):
            sheet[f'A{index}'] = x
            sheet[f'B{index}'] = y
        # 保存 Excel 文件
        workbook.save('obstacle_data.xlsx')

    def load_obs_from_excel(self, obs):
        # obs_last = set()
        workbook = openpyxl.load_workbook("obstacle_data.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True, min_row=2):  # 从第二行开始读取数据，跳过表头
            if len(row) == 2:
                x, y = row
                obs.add((x, y))
        return obs