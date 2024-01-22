"""
Env 2D
@author: huiming zhou
"""

import random
import openpyxl
from create_map import get_map_data

class Env:
    def __init__(self, regenerate=False, use_benchmark=False):
        if use_benchmark:
            self.y_range, self.x_range, self.obs = get_map_data()
            self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1)]  # 4 neighborhoods
            # self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1), (-1, 1), (1, 1), (-1, -1), (-1, 1)]  # 8 neighborhoods
        else:
            self.regenerate_obs = regenerate
            self.x_range = 51  # size of background
            self.y_range = 31
            self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1)] # 4 neighborhoods
            # self.motions = [(-1, 0), (0, 1), (1, 0), (0, -1), (-1, 1), (1, 1), (-1, -1), (-1, 1)]  # 8 neighborhoods
            self.obs = self.obs_map()

    def update_obs(self, obs):
        self.obs = obs

    def obs_map(self):
        """
        Initialize obstacles' positions
        :return: map of obstacles
        """

        x = self.x_range
        y = self.y_range
        obs = set()

        for i in range(x):
            obs.add((i, 0))
        for i in range(x):
            obs.add((i, y - 1))

        for i in range(y):
            obs.add((0, i))
        for i in range(y):
            obs.add((x - 1, i))

        # 随机生成一系列障碍物
        if self.regenerate_obs:
            proportion = 0.2
            num_obstacles = int(proportion * (x - 2) * (y - 2))
            obstacle_data = []  # 存储障碍物坐标信息

            for _ in range(num_obstacles):
                rand_x = random.randint(1, x - 2)  # 排除边界
                rand_y = random.randint(1, y - 2)  # 排除边界
                obs.add((rand_x, rand_y))
                obstacle_data.append((rand_x, rand_y))
                self.save_obstacle_data_to_excel(obstacle_data)

            print("Generate obstacle data successfully")
            return obs
        return self.load_obs_from_excel(obs)

    def save_obstacle_data_to_excel(self, data):
        # 创建一个新的 Excel 文件
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # 添加表头
        sheet['A1'] = 'X Coordinate'
        sheet['B1'] = 'Y Coordinate'

        # 写入障碍物数据
        for index, (x, y) in enumerate(data, start=2):
            sheet[f'A{index}'] = x
            sheet[f'B{index}'] = y

        # 保存 Excel 文件
        workbook.save('obstacle_data.xlsx')

    def load_obs_from_excel(self, obs):
        # obs_last = set()
        workbook = openpyxl.load_workbook("obstacle_data.xlsx")
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True, min_row=2):  # 从第二行开始读取数据，跳过表头
            if len(row) == 2:
                x, y = row
                obs.add((x, y))
        return obs

# 创建 Env 类的实例
# Env_regen = Env(regenerate=True)

